from __future__ import annotations

import csv
import json
import re
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class StepImportError(ValueError):
    pass


SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
BOOL_KEYS = {"clear_first"}
INT_KEYS = {"ms", "timeout_ms", "amount", "duration"}
FLOAT_KEYS = {"seconds", "threshold"}
JSON_COLUMNS = {"step_json", "json"}
STEP_TEXT_COLUMNS = {"step", "instruction", "action_text"}
META_COLUMNS = {"run_name", "test_case_name", "description", "notes", "comment"}


def parse_step_rows_from_upload(filename: str, content: bytes) -> list[dict[str, Any] | str]:
    extension = Path(filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise StepImportError(f"Unsupported file type '{extension}'. Supported: {supported}")
    if not content:
        raise StepImportError("Uploaded file is empty")

    if extension == ".csv":
        rows = _read_csv_rows(content)
    else:
        rows = _read_xlsx_rows(content)

    parsed_steps: list[dict[str, Any] | str] = []
    for row in rows:
        normalized = _normalize_row(row)
        if not normalized:
            continue

        maybe_step = _extract_step_candidate(normalized)
        if maybe_step is not None:
            parsed_steps.append(maybe_step)

    if not parsed_steps:
        raise StepImportError("No step rows found in uploaded file")
    return parsed_steps


def _read_csv_rows(content: bytes) -> list[dict[str, Any]]:
    text = _decode_text(content)
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise StepImportError("CSV file must include a header row")

    rows: list[dict[str, Any]] = []
    for row in reader:
        if not row:
            continue
        rows.append({str(key): value for key, value in row.items() if key is not None})
    return rows


def _read_xlsx_rows(content: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active

    header: list[str] | None = None
    rows: list[dict[str, Any]] = []

    for row_values in sheet.iter_rows(values_only=True):
        if not row_values:
            continue
        if header is None:
            candidate = [str(cell).strip() if cell is not None else "" for cell in row_values]
            if not any(candidate):
                continue
            header = candidate
            continue

        if not any(cell is not None and str(cell).strip() for cell in row_values):
            continue

        row_dict: dict[str, Any] = {}
        for index, key in enumerate(header):
            if not key:
                continue
            value = row_values[index] if index < len(row_values) else None
            row_dict[key] = value
        if row_dict:
            rows.append(row_dict)

    workbook.close()
    if header is None:
        raise StepImportError("Excel file must include a header row")
    return rows


def _decode_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise StepImportError("Could not decode CSV file text")


def _normalize_row(raw_row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for raw_key, raw_value in raw_row.items():
        key = _normalize_key(raw_key)
        if not key:
            continue

        value = _normalize_cell(raw_value)
        if value is None:
            continue

        if key in META_COLUMNS:
            continue
        normalized[key] = _coerce_value(key, value)
    return normalized


def _normalize_key(raw_key: Any) -> str:
    if raw_key is None:
        return ""
    key = str(raw_key).strip().lower()
    if not key:
        return ""
    key = re.sub(r"[^a-z0-9]+", "_", key)
    return key.strip("_")


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text if text else None
    return value


def _coerce_value(key: str, value: Any) -> Any:
    if isinstance(value, str):
        if key in BOOL_KEYS:
            parsed_bool = _parse_bool(value)
            if parsed_bool is not None:
                return parsed_bool
        if key in INT_KEYS:
            parsed_int = _parse_int(value)
            if parsed_int is not None:
                return parsed_int
        if key in FLOAT_KEYS:
            parsed_float = _parse_float(value)
            if parsed_float is not None:
                return parsed_float
    return value


def _extract_step_candidate(row: dict[str, Any]) -> dict[str, Any] | str | None:
    for key in JSON_COLUMNS:
        if key in row:
            raw_json = row[key]
            if isinstance(raw_json, str):
                parsed = _parse_json_step(raw_json)
                if parsed is not None:
                    return parsed
            return None

    for key in STEP_TEXT_COLUMNS:
        if key in row and len(row) == 1:
            text_value = str(row[key]).strip()
            if text_value:
                return text_value

    if "action_type" in row and "type" not in row:
        row["type"] = row.pop("action_type")
    if "expected_value" in row and "value" not in row:
        row["value"] = row.pop("expected_value")

    return row if row else None


def _parse_json_step(raw: str) -> dict[str, Any] | str | None:
    text = raw.strip()
    if not text:
        return None
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError as exc:
        raise StepImportError(f"Invalid JSON step payload: {exc}") from exc

    if isinstance(parsed, dict):
        return parsed
    if isinstance(parsed, str):
        value = parsed.strip()
        return value or None
    raise StepImportError("JSON step payload must be an object or a string")


def _parse_bool(value: str) -> bool | None:
    normalized = value.strip().lower()
    if normalized in {"1", "true", "yes", "y"}:
        return True
    if normalized in {"0", "false", "no", "n"}:
        return False
    return None


def _parse_int(value: str) -> int | None:
    try:
        return int(float(value))
    except ValueError:
        return None


def _parse_float(value: str) -> float | None:
    try:
        return float(value)
    except ValueError:
        return None
